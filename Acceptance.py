import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import datetime

# Define the colors for highlights
grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
red_fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')

# Streamlit App Title
st.title("Report Automator")

# File Uploader
uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file is not None:
    # Load the uploaded file into a DataFrame
    df = pd.read_excel(uploaded_file)

    # Display the raw data
    st.subheader("Uploaded Data Preview")
    st.write(df)

    # Step 2: Add Aging column
    df['Aging'] = (pd.Timestamp('today').normalize() - pd.to_datetime(df['Last Update Date'])).dt.days

    # Step 3: Add Category column
    def categorize_aging(aging):
        if aging <= 2:
            return 'Normal'
        elif 3 <= aging <= 5:
            return 'Overdue'
        elif 6 <= aging <= 8:
            return 'Highrisk'
        elif 9 <= aging <= 11:
            return 'Critical'
        else:
            return 'Alarming'

    df['Category'] = df['Aging'].apply(categorize_aging)

    # Step 1: Rearrange columns
    columns_order = [
        'Current Handler',
        'Last Update Date',
        'Aging',
        'Category',
        'Supplier Name'
    ] + [col for col in df.columns if col not in ['Current Handler', 'Last Update Date', 'Aging', 'Category', 'Supplier Name']]
    df = df[columns_order]

    # Save the updated DataFrame to Excel
    output_path = "Processed_File.xlsx"
    df.to_excel(output_path, index=False, engine='openpyxl')

    # Step 4: Highlight specific columns
    wb = load_workbook(output_path)
    sheet = wb.active

    # Highlight fill color for headings
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Apply highlight to specific columns in the header
    highlight_columns = ['Current Handler', 'Last Update Date', 'Aging', 'Category']
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in highlight_columns:
            sheet.cell(row=1, column=col_idx).fill = yellow_fill

    # Step 5: Create Pivot Table
    pivot_data = df.groupby(['Current Handler', 'Category']).size().unstack(fill_value=0)
    pivot_data['Grand Total'] = pivot_data.sum(axis=1)
    pivot_data.loc['Grand Total'] = pivot_data.sum()

    # Add formatted sheet
    overview_sheet_name = 'Acceptance Overview'
    wb.create_sheet(title=overview_sheet_name)
    overview_sheet = wb[overview_sheet_name]

    # Get today's date
    today_date = datetime.today().strftime('%d-%b-%y')  # Format as '14-Jan-25'

    # Write header row
    headers = [
        f"PAK REP OFFICE ACCEPTANCE OVERVIEW {today_date}",
        "",
        "Current",
        "1.Normal",
        "2.Overdue",
        "3.Highrisk",
        "4.Critical",
        "Alarming",
        "Grand"
    ]
    sub_headers = [
        "Handler",  # "Handler" stays in the first column
        "(0-2) Days",  # Under "Normal"
        "(3-5) Days",  # Under "Overdue"
        "(6-8) Days",  # Under "Highrisk"
        "(9-11) Days",  # Under "Critical"
        "(12-XX) Days",  # Under "Alarming"
        "Total"  # Grand Total
    ]

    # Apply formatting
    font_bold = Font(bold=True)
    alignment_center = Alignment(horizontal='center')

    # Merge and write main header
    overview_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    overview_sheet.cell(row=1, column=1, value=headers[0]).font = font_bold
    overview_sheet.cell(row=1, column=1).alignment = alignment_center

    # Write sub-headers
    for col_idx, header in enumerate(headers[2:], start=1):
        overview_sheet.cell(row=2, column=col_idx, value=header).font = font_bold
        overview_sheet.cell(row=2, column=col_idx).alignment = alignment_center

    for col_idx, sub_header in enumerate(sub_headers, start=1):
        overview_sheet.cell(row=3, column=col_idx, value=sub_header).font = font_bold
        overview_sheet.cell(row=3, column=col_idx).alignment = alignment_center

    # Highlight "PAK REP OFFICE ACCEPTANCE OVERVIEW 13-JAN-25" with grey
    overview_sheet.cell(row=1, column=1).fill = grey_fill  # Grey fill for the header

    # Highlight "Current Handler" with blue
    overview_sheet.cell(row=3, column=1).fill = blue_fill

    # Highlight "1.Normal (0-2) Days" with green
    overview_sheet.cell(row=3, column=3).fill = green_fill

    # Highlight "2.Overdue, 3.Highrisk, 4.Critical, Alarming" with red
    for col_idx in range(4, 8):  # Columns for these categories (Overdue, Highrisk, Critical, Alarming)
        overview_sheet.cell(row=3, column=col_idx).fill = red_fill

    # Highlight "Grand Total" with blue
    overview_sheet.cell(row=3, column=len(sub_headers)).fill = blue_fill

    # Write pivot table data
    aging_categories = ['Normal', 'Overdue', 'Highrisk', 'Critical', 'Alarming']
    for row_idx, (handler, row) in enumerate(pivot_data.iterrows(), start=4):
        overview_sheet.cell(row=row_idx, column=1, value=handler)
        for col_idx, category in enumerate(aging_categories, start=2):
            overview_sheet.cell(row=row_idx, column=col_idx, value=row.get(category, 0))
        overview_sheet.cell(row=row_idx, column=len(aging_categories) + 2, value=row['Grand Total'])

    # Save the workbook
    wb.save(output_path)

    # Provide download link for the processed file
    st.subheader("Download Processed File")
    with open(output_path, "rb") as file:
        st.download_button(
            label="Download Processed Excel File",
            data=file,
            file_name="Processed_File.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.success("File processed successfully!")